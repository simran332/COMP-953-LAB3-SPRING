import os
import sys
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, numbers

def main():
    sales_csv = get_sales_csv()
    orders_dir = create_orders_dir(sales_csv)
    process_sales_data(sales_csv, orders_dir)

# Get path of sales data CSV file from the command line
def get_sales_csv():
    # Check whether command line parameter provided
    if len(sys.argv) < 2:
        print('Error: Missing the parameter for CSV filepath')
        sys.exit(1)

    # Check whether provided parameter is a valid path of file
    sales_csv_path = sys.argv[1]
    if not os.path.isfile(sales_csv_path):
        print('Error: Provided file does not exist in the CSV filepath')
        sys.exit(1)
    return sales_csv_path

# Create the directory to hold the individual order Excel sheets
def create_orders_dir(sales_csv):
    # Get directory in which sales data CSV file resides
    sales_dir_path = os.path.dirname(os.path.abspath(sales_csv))
    # Determine the name and path of the directory to hold the order data files
    todays_date = date.today().isoformat()
    order_dir_name = f'orders_{todays_date}'
    orders_dir_path = os.path.join(sales_dir_path, order_dir_name)

    # Create the order directory if it does not already exist
    if not os.path.isdir(orders_dir_path):
        os.makedirs(orders_dir_path)

    return orders_dir_path 

# Split the sales data into individual orders and save to Excel sheets
def process_sales_data(sales_csv, orders_dir):
    # Import the sales data from the CSV file into a DataFrame
    orders_dataframe = pd.read_csv(sales_csv)

    # Insert a new "TOTAL PRICE" column into the DataFrame
    orders_dataframe["TOTAL PRICE"] = orders_dataframe["ITEM QUANTITY"] * orders_dataframe["ITEM PRICE"]

    # Remove columns from the DataFrame that are not needed
    orders_dataframe.drop(columns=["ADDRESS", "CITY", "STATE", "POSTAL CODE", "COUNTRY"], inplace=True)

    # Group the rows in the DataFrame by order ID
    grouped = orders_dataframe.groupby("ORDER ID")
    for order_id, order_df in grouped:
        # For each order ID:
        # Remove the "ORDER ID" column
        order_df = order_df.drop(columns=["ORDER ID"])
        # Sort the items by item number
        order_df = order_df.sort_values(by="ITEM NUMBER")
        
        # Append a "GRAND TOTAL" row
        grand_total = order_df["TOTAL PRICE"].sum()
        grand_total_df = pd.DataFrame({'ITEM NUMBER': ['GRAND TOTAL:'], 'TOTAL PRICE': [grand_total]})
        order_df = pd.concat([order_df, grand_total_df], ignore_index=True)

        # Determine the file name and full path of the Excel sheet
        file_name = f'order_{order_id}.xlsx'
        file_path = os.path.join(orders_dir, file_name)

        # Export the data to an Excel sheet
        sheet_name = f'Order_{order_id}'
        order_df.to_excel(file_path, index=False, sheet_name=sheet_name)

        # Format the Excel sheet
        format_excel_sheet(file_path, sheet_name)

# Function to format the Excel sheet
def format_excel_sheet(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Define format for the money columns
    money_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Set column widths and apply formatting
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            # Apply currency format to TOTAL PRICE column
            if cell.column_letter in ['F', 'G'] and cell.row > 1:
                cell.number_format = money_format

    # Save the changes
    wb.save(file_path)

if __name__ == "__main__":
    main()
